import pandas as pd
import os
import glob
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, numbers

def load_abbreviations():
    """
    加载简称/缩写对照表
    
    返回:
    dict: 全称到简称/缩写的映射字典
    """
    download_path = os.path.expanduser("~/Downloads")
    abbr_file = os.path.join(download_path, "简称.xlsx")
    
    if not os.path.exists(abbr_file):
        print(f"警告: 对照表文件 {abbr_file} 不存在")
        return {}
    
    try:
        abbr_df = pd.read_excel(abbr_file)
        if len(abbr_df.columns) < 2:
            print("警告: 对照表格式不正确，需要至少两列数据")
            return {}
            
        # 检查列名
        if '广告主' in abbr_df.columns and '简称' in abbr_df.columns:
            # 使用正确的列名
            abbr_dict = dict(zip(abbr_df['广告主'], abbr_df['简称']))
        else:
            # 使用第一列和第二列，无论列名是什么
            full_name_col = abbr_df.columns[0]
            abbr_col = abbr_df.columns[1]
            abbr_dict = dict(zip(abbr_df[full_name_col], abbr_df[abbr_col]))
            print(f"注意: 使用列 '{full_name_col}' 作为全称，'{abbr_col}' 作为简称/缩写")
        
        print(f"已加载 {len(abbr_dict)} 个对照项")
        
        # 打印一些样本，用于验证
        sample_count = min(5, len(abbr_dict))
        if sample_count > 0:
            print("对照表样本:")
            for i, (full, abbr) in enumerate(list(abbr_dict.items())[:sample_count]):
                print(f"  {i+1}. {full} -> {abbr}")
        
        return abbr_dict
    except Exception as e:
        print(f"加载对照表时出错: {str(e)}")
        return {}
   

def process_ad_data(input_date, abbr_dict=None):
    """
    处理广告数据，根据日期筛选，按广告主分组计算曝光和税后收入
    
    参数:
    input_date (str): 需要筛选的日期，格式为 YYYY-MM-DD
    abbr_dict (dict): 全称到简称/缩写的映射字典
    
    返回:
    tuple: 包含两个DataFrame的元组，第一个是所有广告主的汇总，第二个是Splash广告位的汇总
    """
    if abbr_dict is None:
        abbr_dict = {}
        
    # 将输入日期转换为datetime对象以便于比较
    target_date = pd.to_datetime(input_date)
    
    # 查找最新的效果日报数据文件
    download_path = os.path.expanduser("~/Downloads")
    pattern = os.path.join(download_path, "效果日报数据*.xlsx")
    files = glob.glob(pattern)
    
    if not files:
        raise FileNotFoundError(f"在 {download_path} 中未找到匹配 '效果日报数据*.xlsx' 的文件")
    
    # 按文件修改时间排序，获取最新的文件
    latest_file = max(files, key=os.path.getmtime)
    print(f"正在处理文件: {latest_file}")
    
    # 读取excel文件中的"收入源数据-当月"表
    df = pd.read_excel(latest_file, sheet_name="收入源数据-当月")
    
    # 确保日期列是datetime类型
    df['日期'] = pd.to_datetime(df['日期'])
    
    # 筛选指定日期的数据
    filtered_df = df[df['日期'] == target_date]
    
    if filtered_df.empty:
        print(f"警告: 在数据中没有找到日期 {input_date} 的记录")
        # 返回空的DataFrame，使用新的列名
        empty_df = pd.DataFrame(columns=['行标签', '求和项:税后收入', '求和项:曝光', '分类', '简称'])
        return empty_df, empty_df
    
    # 表1: 按广告主分组，计算曝光和税后收入的总和
    table1 = filtered_df.groupby('广告主').agg({
        '税后收入': 'sum',
        '曝光': 'sum'
    }).reset_index()
    
    # 重命名列以匹配目标Excel文件的格式
    table1 = table1.rename(columns={
        '广告主': '行标签',
        '税后收入': '求和项:税后收入',
        '曝光': '求和项:曝光'
    })
    
    # 在表1的最前面添加日期列，使用yyyy/m/d格式
    formatted_date = f"{target_date.year}/{target_date.month}/{target_date.day}"
    table1.insert(0, '日期', formatted_date)
    
    # 添加分类列，值为"外包"
    table1['分类'] = '外包'
    
    # 添加简称列
    table1['简称'] = table1['行标签'].map(lambda x: abbr_dict.get(x, ''))
    
    # 检查未匹配的行标签
    missing_abbr = table1[table1['简称'] == '']['行标签'].unique()
    if len(missing_abbr) > 0:
        print("\n警告: 以下广告主没有对应的简称:")
        for name in missing_abbr:
            print(f"  - {name}")
    
    # 表2: 筛选广告位包含"Splash"或"splash"的数据
    splash_filter = filtered_df['广告位'].str.contains('Splash|splash', regex=True, na=False)
    splash_df = filtered_df[splash_filter]
    
    # 按广告主分组，计算曝光和税后收入的总和
    table2 = splash_df.groupby('广告主').agg({
        '税后收入': 'sum',
        '曝光': 'sum'
    }).reset_index()
    
    # 重命名列以匹配目标Excel文件的格式
    table2 = table2.rename(columns={
        '广告主': '行标签',
        '税后收入': '求和项:税后收入',
        '曝光': '求和项:曝光'
    })
    
    # 在表2的最前面添加日期列，使用yyyy/m/d格式
    table2.insert(0, '日期', formatted_date)
    
    # 添加分类列，值为"外包"
    table2['分类'] = '外包'
    
    # 添加简称列
    table2['简称'] = table2['行标签'].map(lambda x: abbr_dict.get(x, ''))
    
    # 检查未匹配的行标签（开屏数据）
    missing_abbr_splash = table2[table2['简称'] == '']['行标签'].unique()
    if len(missing_abbr_splash) > 0:
        print("\n警告: 以下开屏广告主没有对应的简称:")
        for name in missing_abbr_splash:
            print(f"  - {name}")
    
    return table1, table2

def save_to_excel(table1, table2, input_date):
    """
    将两个表格保存到一个Excel文件中的两个不同工作表
    
    参数:
    table1 (DataFrame): 第一个表格，所有广告主的汇总
    table2 (DataFrame): 第二个表格，Splash广告位的汇总
    input_date (str): 筛选日期，用于文件命名
    
    返回:
    str: 保存的文件路径
    """
    # 使用yyyy-mm-dd格式创建输出文件名，避免路径中的斜杠问题
    formatted_date = pd.to_datetime(input_date).strftime('%Y-%m-%d')
    output_file = os.path.expanduser(f"~/Downloads/【数透】广告数据分析_{formatted_date}.xlsx")
    
    # 创建ExcelWriter对象
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 将表1写入'所有广告主'工作表
        table1.to_excel(writer, sheet_name='所有广告主', index=False)
        
        # 将表2写入'Splash广告位'工作表
        table2.to_excel(writer, sheet_name='Splash广告位', index=False)
    
    print(f"数据已保存到: {output_file}")
    return output_file

def append_to_existing_excel(table1, table2):
    """
    将处理结果追加到指定的Excel文件的特定工作表中
    
    参数:
    table1 (DataFrame): 所有广告主的汇总数据
    table2 (DataFrame): Splash广告位的汇总数据
    """
    target_file = os.path.expanduser("~/Downloads/外包自营.xlsx")
    
    if not os.path.exists(target_file):
        print(f"警告: 目标文件 {target_file} 不存在，将创建新文件")
        # 创建新文件
        with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
            # 创建空的工作表
            pd.DataFrame().to_excel(writer, sheet_name='总', index=False)
            pd.DataFrame().to_excel(writer, sheet_name='开屏', index=False)
    
    try:
        # 读取现有文件中的数据
        existing_total = pd.read_excel(target_file, sheet_name='总')
        existing_splash = pd.read_excel(target_file, sheet_name='开屏')
        
        # 检查是否需要重命名简称列
        if '缩写' in existing_total.columns and '简称' in table1.columns:
            table1 = table1.rename(columns={'简称': '缩写'})
        if '缩写' in existing_splash.columns and '简称' in table2.columns:
            table2 = table2.rename(columns={'简称': '缩写'})
            
        # 确保两个DataFrame具有相同的列
        for col in existing_total.columns:
            if col not in table1.columns:
                table1[col] = ''
        
        for col in existing_splash.columns:
            if col not in table2.columns:
                table2[col] = ''
        
        # 将新数据与现有数据合并
        updated_total = pd.concat([existing_total, table1[existing_total.columns]], ignore_index=True)
        updated_splash = pd.concat([existing_splash, table2[existing_splash.columns]], ignore_index=True)
        
        # 按日期排序
        if '日期' in updated_total.columns:
            updated_total['日期_排序'] = pd.to_datetime(updated_total['日期'])
            updated_total = updated_total.sort_values('日期_排序')
            updated_total = updated_total.drop(columns=['日期_排序'])
        
        if '日期' in updated_splash.columns:
            updated_splash['日期_排序'] = pd.to_datetime(updated_splash['日期'])
            updated_splash = updated_splash.sort_values('日期_排序')
            updated_splash = updated_splash.drop(columns=['日期_排序'])
        
        # 保存更新后的数据
        with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
            updated_total.to_excel(writer, sheet_name='总', index=False)
            updated_splash.to_excel(writer, sheet_name='开屏', index=False)
        
        print(f"数据已成功追加到 {target_file}")
        print(f"  - '总' 工作表: {len(table1)} 行新数据, 共 {len(updated_total)} 行")
        print(f"  - '开屏' 工作表: {len(table2)} 行新数据, 共 {len(updated_splash)} 行")
        
    except Exception as e:
        print(f"追加数据到既有Excel文件时出错: {str(e)}")
        raise

def fix_date_format_in_excel(file_path):
    """
    修复Excel文件中日期格式，将其统一为yyyy/m/d格式
    
    参数:
    file_path (str): Excel文件路径
    """
    try:
        print(f"\n开始修复 {file_path} 中的日期格式...")
        
        # 创建日期格式
        date_format = NamedStyle(name='yyyy/m/d')
        date_format.number_format = 'yyyy/m/d'
        
        # 加载工作簿
        book = load_workbook(file_path)
        
        # 注册样式
        if date_format.name not in book.named_styles:
            book.add_named_style(date_format)
        
        # 处理每个工作表
        for sheet_name in ['总', '开屏']:
            if sheet_name not in book.sheetnames:
                print(f"  - 工作表 {sheet_name} 不存在，已跳过")
                continue
                
            ws = book[sheet_name]
            
            # 检查是否有数据
            if ws.max_row <= 1:
                print(f"  - 工作表 {sheet_name} 为空或只有标题行，已跳过")
                continue
            
            # 获取日期列的索引（假设是第一列）
            date_col_idx = 1
            
            # 获取列标题
            first_cell = ws.cell(row=1, column=date_col_idx)
            first_col_header = first_cell.value
            
            print(f"  - 修复工作表 {sheet_name} 中的 '{first_col_header}' 列")
            
            # 转换日期格式
            converted_count = 0
            for row in range(2, ws.max_row + 1):  # 跳过标题行
                cell = ws.cell(row=row, column=date_col_idx)
                if cell.value and cell.value != 'N/A':
                    try:
                        # 转换为日期对象
                        if isinstance(cell.value, datetime):
                            date_obj = cell.value
                        else:
                            # 尝试解析字符串日期
                            date_str = str(cell.value).strip()
                            date_obj = pd.to_datetime(date_str)
                        
                        # 设置日期格式
                        cell.value = date_obj
                        cell.number_format = 'yyyy/m/d'
                        converted_count += 1
                    except:
                        pass  # 如果无法转换，保持原值
            
            print(f"  - 工作表 {sheet_name} 中已转换 {converted_count} 个日期")
        
        # 保存工作簿
        book.save(file_path)
        print(f"✓ 日期格式修复完成，已保存文件")
        
    except Exception as e:
        print(f"修复日期格式时出错: {str(e)}")
        import traceback
        print(traceback.format_exc())

def preview_data(table1, table2):
    """
    预览两个表格的数据
    
    参数:
    table1 (DataFrame): 第一个表格
    table2 (DataFrame): 第二个表格
    """
    print("\n所有广告主数据预览:")
    print(table1)
    
    print("\nSplash广告位数据预览:")
    print(table2)

def main():
    # 获取前一天的日期作为默认值
    yesterday = (datetime.now() - pd.Timedelta(days=1)).strftime('%Y-%m-%d')
    print(f"默认日期为前一天: {yesterday}")
    
    # 加载简称对照表
    abbr_dict = load_abbreviations()
    
    # 询问是否需要调整日期
    adjust_date = input("是否需要调整日期? (y/n，默认n): ").lower()
    
    if adjust_date == 'y':
        # 获取用户输入的mmdd格式日期
        while True:
            input_date = input("请输入要筛选的日期 (格式: MMDD): ")
            # 验证MMDD格式
            date_pattern = re.compile(r'^\d{4}$')
            if date_pattern.match(input_date):
                try:
                    # 添加当前年份，并转换为标准日期格式
                    current_year = datetime.now().year
                    month = int(input_date[:2])
                    day = int(input_date[2:])
                    
                    # 验证月份和日期是否有效
                    if 1 <= month <= 12 and 1 <= day <= 31:
                        full_date = f"{current_year}-{month:02d}-{day:02d}"
                        # 尝试转换为日期对象来验证日期是否有效
                        datetime.strptime(full_date, '%Y-%m-%d')
                        input_date = full_date
                        break
                    else:
                        print("无效的月份或日期，请重新输入。")
                except ValueError:
                    print("无效的日期，请输入正确的日期。")
            else:
                print("日期格式不正确，请使用MMDD格式。")
    else:
        input_date = yesterday
    
    try:
        # 处理数据，传入简称字典
        table1, table2 = process_ad_data(input_date, abbr_dict)
        
        # 预览数据
        preview_data(table1, table2)
        
        # 保存数据到单独文件
        output_file = save_to_excel(table1, table2, input_date)
        
        # 追加数据到外包自营.xlsx文件
        append_to_existing_excel(table1, table2)
        
        # 修复外包自营.xlsx中的日期格式
        target_file = os.path.expanduser("~/Downloads/外包自营.xlsx")
        fix_date_format_in_excel(target_file)
        
        print(f"\n处理完成！数据已保存到: {output_file}")
        print(f"并已追加到: {target_file}")
        
    except Exception as e:
        print(f"处理数据时发生错误: {str(e)}")

if __name__ == "__main__":
    main()
